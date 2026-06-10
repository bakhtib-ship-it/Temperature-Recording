import io
import os
import threading
import webbrowser
from datetime import date, timedelta
import requests
from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Every city paired with its latitude and longitude.
# Latitude is how far north/south a place is; longitude is how far east/west.
# Open-Meteo uses these coordinates to find the right weather station.
CITIES = {
    "Asia": [
        {"name": "Tokyo, Japan",        "lat": 35.6762,  "lon": 139.6503},
        {"name": "Delhi, India",         "lat": 28.7041,  "lon": 77.1025},
        {"name": "Shanghai, China",      "lat": 31.2304,  "lon": 121.4737},
        {"name": "Dhaka, Bangladesh",    "lat": 23.8103,  "lon": 90.4125},
        {"name": "Beijing, China",       "lat": 39.9042,  "lon": 116.4074},
        {"name": "Mumbai, India",        "lat": 19.0760,  "lon": 72.8777},
        {"name": "Osaka, Japan",         "lat": 34.6937,  "lon": 135.5023},
        {"name": "Karachi, Pakistan",    "lat": 24.8607,  "lon": 67.0011},
        {"name": "Chongqing, China",     "lat": 29.4316,  "lon": 106.9123},
        {"name": "Istanbul, Turkey",     "lat": 41.0082,  "lon": 28.9784},
    ],
    "South & Central America": [
        {"name": "São Paulo, Brazil",       "lat": -23.5505, "lon": -46.6333},
        {"name": "Mexico City, Mexico",     "lat": 19.4326,  "lon": -99.1332},
        {"name": "Buenos Aires, Argentina", "lat": -34.6037, "lon": -58.3816},
        {"name": "Rio de Janeiro, Brazil",  "lat": -22.9068, "lon": -43.1729},
        {"name": "Lima, Peru",              "lat": -12.0464, "lon": -77.0428},
        {"name": "Bogotá, Colombia",        "lat": 4.7110,   "lon": -74.0721},
    ],
    "Africa": [
        {"name": "Cairo, Egypt",               "lat": 30.0444,  "lon": 31.2357},
        {"name": "Lagos, Nigeria",             "lat": 6.5244,   "lon": 3.3792},
        {"name": "Kinshasa, DR Congo",         "lat": -4.4419,  "lon": 15.2663},
        {"name": "Johannesburg, South Africa", "lat": -26.2041, "lon": 28.0473},
    ],
    "North America": [
        {"name": "New York City, USA",   "lat": 40.7128,  "lon": -74.0060},
        {"name": "Los Angeles, USA",     "lat": 34.0522,  "lon": -118.2437},
        {"name": "Chicago, USA",         "lat": 41.8781,  "lon": -87.6298},
        {"name": "Dallas-Fort Worth, USA","lat": 32.7767, "lon": -96.7970},
        {"name": "Toronto, Canada",      "lat": 43.6532,  "lon": -79.3832},
        {"name": "Houston, USA",         "lat": 29.7604,  "lon": -95.3698},
        {"name": "Washington D.C., USA", "lat": 38.9072,  "lon": -77.0369},
        {"name": "Miami, USA",           "lat": 25.7617,  "lon": -80.1918},
        {"name": "Atlanta, USA",         "lat": 33.7490,  "lon": -84.3880},
        {"name": "Philadelphia, USA",    "lat": 39.9526,  "lon": -75.1652},
        {"name": "Vancouver, Canada",    "lat": 49.2827,  "lon": -123.1207},
        {"name": "Montreal, Canada",     "lat": 45.5017,  "lon": -73.5673},
    ],
    "Europe": [
        {"name": "Moscow, Russia",    "lat": 55.7558, "lon": 37.6173},
        {"name": "London, UK",        "lat": 51.5074, "lon": -0.1278},
        {"name": "Paris, France",     "lat": 48.8566, "lon": 2.3522},
        {"name": "Madrid, Spain",     "lat": 40.4168, "lon": -3.7038},
        {"name": "Barcelona, Spain",  "lat": 41.3851, "lon": 2.1734},
        {"name": "Berlin, Germany",   "lat": 52.5200, "lon": 13.4050},
        {"name": "Rome, Italy",       "lat": 41.9028, "lon": 12.4964},
    ],
    "Middle East & Oceania": [
        {"name": "Tehran, Iran",           "lat": 35.6892,  "lon": 51.3890},
        {"name": "Riyadh, Saudi Arabia",   "lat": 24.7136,  "lon": 46.6753},
        {"name": "Sydney, Australia",      "lat": -33.8688, "lon": 151.2093},
        {"name": "Melbourne, Australia",   "lat": -37.8136, "lon": 144.9631},
    ],
}


def build_city_lookup():
    """Return a flat dict of {city_name: (lat, lon)} for quick lookup."""
    lookup = {}
    for city_list in CITIES.values():
        for city in city_list:
            lookup[city["name"]] = (city["lat"], city["lon"])
    return lookup


def fetch_temperatures_30_days(lat, lon):
    """Ask Open-Meteo's historical archive for the past 30 days of daily mean
    temperatures at a given location. Returns a list of (date_string, temp) pairs,
    e.g. [("2026-05-02", 18.4), ("2026-05-03", 19.1), ...]."""
    end_date   = date.today() - timedelta(days=1)   # yesterday — today not archived yet
    start_date = end_date - timedelta(days=29)       # 30 days total
    url = (
        "https://archive-api.open-meteo.com/v1/archive"
        f"?latitude={lat}&longitude={lon}"
        f"&start_date={start_date}&end_date={end_date}"
        "&daily=temperature_2m_mean&timezone=UTC"
    )
    resp = requests.get(url, timeout=15)
    resp.raise_for_status()
    data = resp.json()["daily"]
    return list(zip(data["time"], data["temperature_2m_mean"]))


def build_excel(results):
    """Create an Excel workbook where rows = cities, columns = dates.
    `results` is a list of (city_name, [(date_str, temp), ...]) tuples."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Temperatures"

    header_fill = PatternFill("solid", fgColor="2E75B6")
    header_font = Font(bold=True, color="FFFFFF")
    center      = Alignment(horizontal="center")

    # Pull the list of dates from the first city — every city has the same dates.
    dates = [entry[0] for entry in results[0][1]]

    # --- Header row: "City" in A1, then one date per column ---
    ws["A1"].value     = "City"
    ws["A1"].font      = header_font
    ws["A1"].fill      = header_fill
    ws["A1"].alignment = center

    for col_idx, d in enumerate(dates, start=2):
        cell           = ws.cell(row=1, column=col_idx, value=d)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    # --- Data rows: one row per city ---
    for row_idx, (city, temp_data) in enumerate(results, start=2):
        # City name in column A
        ws.cell(row=row_idx, column=1, value=city)

        # One temperature per date column
        for col_idx, (_, temp) in enumerate(temp_data, start=2):
            cell           = ws.cell(row=row_idx, column=col_idx, value=temp)
            cell.alignment = center

        # Alternate row shading across the full width of the table
        row_fill = PatternFill("solid", fgColor="D6E4F0" if row_idx % 2 == 0 else "EBF3FB")
        for col_idx in range(1, len(dates) + 2):
            ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # --- Column widths ---
    ws.column_dimensions["A"].width = 28          # city names
    for col_idx in range(2, len(dates) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13   # date columns

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/")
def index():
    return render_template("index.html", cities=CITIES)


@app.route("/download", methods=["POST"])
def download():
    selected_names = request.json.get("cities", [])

    if not selected_names:
        return {"error": "No cities selected."}, 400

    city_lookup = build_city_lookup()
    results = []

    for name in selected_names:
        if name in city_lookup:
            lat, lon = city_lookup[name]
            temp_series = fetch_temperatures_30_days(lat, lon)
            results.append((name, temp_series))

    excel_buf = build_excel(results)

    return send_file(
        excel_buf,
        download_name="temperatures.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    threading.Timer(1.0, lambda: webbrowser.open("http://127.0.0.1:5000")).start()
    app.run(debug=True, use_reloader=False)
